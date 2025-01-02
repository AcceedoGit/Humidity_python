from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import FileResponse
from fastapi.responses import JSONResponse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import pytz 
import os 
from fpdf import FPDF
from configuration.database import Board_1, Board_2, Board_3,db
from statistics import mean
from typing import Dict


# Initialize FastAPI and Router
app = FastAPI()
ReportRouter = APIRouter()

# MongoDB Collections Mapping
BOARD_COLLECTIONS = {
    1: Board_1,
    2: Board_2,
    3: Board_3,
}

# Timezone and Directory Setup
IST = pytz.timezone('Asia/Kolkata')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMP_DIR = os.path.join(BASE_DIR, "excel_temp")
IMAGE_DIR = os.path.join(BASE_DIR, "graph_images")
os.makedirs(TEMP_DIR, exist_ok=True)
os.makedirs(IMAGE_DIR, exist_ok=True)

# Function to query MongoDB data
def query_data(unit_ID: int):
    if unit_ID not in BOARD_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Invalid unit ID")
    
    collection = BOARD_COLLECTIONS[unit_ID]
    now = datetime.now(IST)
    start_dt = now.replace(hour=8, minute=30, second=0, microsecond=0)
    end_dt = (start_dt + timedelta(days=1)).replace(hour=8, minute=29, second=59)

    data = collection.find({
        "created_at": {"$gte": start_dt, "$lt": end_dt}
    }).sort("created_at", 1)
    return list(data)

# Function to generate the graph
def generate_graph(times, temperatures, humidities, unit_ID):
    plt.switch_backend('Agg')  # Use non-interactive backend
    plt.figure(figsize=(10, 5))
    plt.plot(times, temperatures, label='Temperature (°C)', color='r')
    plt.plot(times, humidities, label='Humidity (%)', color='b')
    plt.title(f'Graph Data for Unit ID: {unit_ID}')
    plt.xlabel('Time (IST)')
    plt.ylabel('Values')
    plt.xticks(rotation=45)
    plt.legend()
    plt.grid()

    # Save graph image
    graph_image_path = f"{IMAGE_DIR}/graph_data_unit_{unit_ID}.png"
    try:
        plt.savefig(graph_image_path)
        plt.close()
        return graph_image_path  # Ensure path is returned
    except Exception as e:
        print(f"Error saving graph: {e}")
        return None

# Excel Generation Endpoint
@ReportRouter.get("/download/excel/{unit_ID}", response_class=FileResponse)
async def download_excel(unit_ID: int):
    data = query_data(unit_ID)
    filename = os.path.join(TEMP_DIR, f"graph_data_unit_{unit_ID}.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Unit {unit_ID} Data"

    headers = ["Time (IST)", "Temperature (°C)", "Humidity (%)"]
    sheet.append(headers)

    times, temperatures, humidities = [], [], []

    for entry in data:
        time = entry["created_at"].astimezone(IST)
        temperature = entry.get("t", 0)
        humidity = entry.get("h", 0)
        sheet.append([time.strftime("%Y-%m-%d %I:%M:%S %p"), temperature, humidity])
        times.append(time)
        temperatures.append(temperature)
        humidities.append(humidity)

    graph_image_path = generate_graph(times, temperatures, humidities, unit_ID)
    img = Image(graph_image_path)
    sheet.add_image(img, 'E5')  # Place image at cell E5
    workbook.save(filename)

    return FileResponse(filename, filename=f"graph_data_unit_{unit_ID}.xlsx")

# PDF Generation Endpoint
@ReportRouter.get("/download/pdf/{unit_ID}", response_class=FileResponse)
async def download_pdf(unit_ID: int):
    data = query_data(unit_ID)
    filename = os.path.join(TEMP_DIR, f"graph_data_unit_{unit_ID}.pdf")

    times, temperatures, humidities = [], [], []

    for entry in data:
        time = entry["created_at"].astimezone(IST)
        temperature = entry.get("t", 0)
        humidity = entry.get("h", 0)
        times.append(time)
        temperatures.append(temperature)
        humidities.append(humidity)

    graph_image_path = generate_graph(times, temperatures, humidities, unit_ID)

    if not graph_image_path or not os.path.exists(graph_image_path):
        raise HTTPException(status_code=500, detail="Graph image could not be generated.")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt=f"Graph Data for Unit {unit_ID}", ln=True, align='C')
    pdf.ln(10)

    pdf.cell(60, 10, txt="Time (IST)", border=1, align='C')
    pdf.cell(60, 10, txt="Temperature (°C)", border=1, align='C')
    pdf.cell(60, 10, txt="Humidity (%)", border=1, align='C')
    pdf.ln(10)

    for t, temp, hum in zip(times, temperatures, humidities):
        pdf.cell(60, 10, txt=t.strftime("%Y-%m-%d %I:%M:%S %p"), border=1, align='C')
        pdf.cell(60, 10, txt=str(temp), border=1, align='C')
        pdf.cell(60, 10, txt=str(hum), border=1, align='C')
        pdf.ln(10)

    # Add the graph image
    pdf.image(graph_image_path, x=10, y=80, w=190)

    pdf.output(filename)
    return FileResponse(filename, filename=f"graph_data_unit_{unit_ID}.pdf")

def get_monthly_avg(unit_ID: int, month: int, year: int):
    collection_name = BOARD_COLLECTIONS.get(unit_ID)
    
    if collection_name is None:
        raise HTTPException(status_code=404, detail="Unit ID not found in the database.")
    
    # Ensure collection_name is a string and not a collection object
    if isinstance(collection_name, str):
        collection = db[collection_name]
    else:
        # If it's a Collection object, get the name using `collection.name`
        collection = collection_name
        collection_name = collection.name
    
    print(f"Using collection: {collection_name}")

    # Date range for the given month and year
    start_date = datetime(year, month, 1)
    end_date = datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1)
    
    data = collection.find({
        "created_at": {"$gte": start_date, "$lt": end_date}
    })
    
    total_temp = total_humidity = count = 0
    for entry in data:
        # Safely handle None values for "t" and "h" by treating them as 0 if None
        temp = entry.get("t", 0) or 0
        humidity = entry.get("h", 0) or 0
        total_temp += temp
        total_humidity += humidity
        count += 1
    
    if count == 0:
        raise HTTPException(status_code=404, detail="No data found for the given month.")
    
    avg_temp = total_temp / count
    avg_humidity = total_humidity / count
    
    return {"unit_ID": unit_ID, "month": month, "year": year, "avg_temp": avg_temp, "avg_humidity": avg_humidity}

# FastAPI endpoint to fetch monthly averages
@ReportRouter.get("/average/{unit_ID}")
async def monthly_average(unit_ID: int, month: int, year: int):
    try:
        result = get_monthly_avg(unit_ID, month, year)
        return JSONResponse(content=result)
    except HTTPException as e:
        raise e
